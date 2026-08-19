import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from aiogram import Router, F, types
from aiogram.filters import CommandStart, Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile
import database as db

router = Router()

MONTH_NAMES = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December"
}

class Form(StatesGroup):
    add_category = State()
    input_pages = State()
    edit_pages = State()
    create_group_name = State()
    create_group_cats = State()
    custom_date_range = State()

def main_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📂 Категориялар"), KeyboardButton(text="📊 Статистика")]
        ],
        resize_keyboard=True
    )

def cancel_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="❌ Отмена")]
        ],
        resize_keyboard=True
    )

# --- РЕСЕТ КОМАНДАСЫ (/reset болуп өзгөртүлдү) ---
# --- РЕСЕТ КОМАНДАСЫ ---
@router.message(Command("reset"))
@router.message(F.text.startswith("/reset"))
async def cmd_reset(message: types.Message, state: FSMContext):
    try:
        await state.clear()
        
        # Колдонуучунун бир гана жеке данныеларын тазалайт
        # (Башка колдонуучуларга да, топторго да тийбейт)
        await db.clear_user_full_data(message.from_user.id)
        
        # Колдонуучуну базага кайра каттоо
        await db.add_user(
            message.from_user.id, 
            message.from_user.username, 
            message.from_user.full_name
        )
        
        await message.answer(
            "🔄 **Сиздин жеке маалыматтарыңыз жана тесттик логдоруңуз тазаланды!**\n"
            "Эми ботту жаңыдан колдонсоңуз болот.",
            parse_mode="Markdown",
            reply_markup=main_kb()
        )
    except Exception as e:
        await message.answer(f"⚠️ **Ката чыкты:**\n`{e}`", parse_mode="Markdown")

@router.message(Command("cancel"))
@router.message(F.text == "❌ Отмена")
async def cmd_cancel(message: types.Message, state: FSMContext):
    await state.clear()
    await message.answer("🔄 **Процесс жокко чыгарылды!**", parse_mode="Markdown", reply_markup=main_kb())

@router.message(Command("rename"))
async def cmd_rename(message: types.Message):
    args = message.text.split(maxsplit=2)
    if len(args) < 3:
        await message.answer("⚠️ Форматы: `/rename <эски_аты> <жаңы_аты>`", parse_mode="Markdown")
        return
    
    old_title = args[1].strip()
    new_title = args[2].strip()
    
    success = await db.rename_category_by_title(message.from_user.id, old_title, new_title)
    if success:
        await message.answer(f"✅ «**{old_title}**» категориясы «**{new_title}**» болуп өзгөртүлдү!", parse_mode="Markdown")
    else:
        await message.answer(f"❌ «**{old_title}**» деген категория табылган жок.", parse_mode="Markdown")

# --- ЖАҢЫЛАНГАН САЛАМДАШУУ ТЕКСТИ ---
@router.message(CommandStart())
async def cmd_start(message: types.Message, state: FSMContext):
    await state.clear()
    await db.add_user(message.from_user.id, message.from_user.username, message.from_user.full_name)
    
    user_name = message.from_user.first_name
    start_text = (
        f"Саламатсызбы, {user_name}! 👋\n\n"
        "📖 **«Book Tracker» ботуна кош келиңиз!**\n\n"
        "Бул бот китеп окуу адаттарыңызды калыптандырууга жана **окуу конкурстарын/мелдештерин** уюштурууга жардам берет.\n\n"
        "👤 **Жөнөкөй колдонуучулар үчүн:**\n"
        "• Күн сайын окуган беттерди киргизүү.\n"
        "• `/pin <код>` — Конкурска же топко кошулуу.\n"
        "• `/rename <эски> <жаңы>` — Категория атын өзгөртүү.\n\n"
        "👑 **Админдер/Уюштуруучулар үчүн:**\n"
        "• `/make_admin` — Жаңы топ жана конкурс түзүү (ПИН-код берет).\n"
        "• `/admin` — Конкурс аралыгындагы датаны тандап, толук Excel отчет алуу.\n\n"
        "⚙️ **Кызматтык командалар:**\n"
        "• `/reset` — Базаны толук тазалоо (сак болуңуз!)."
    )
    await message.answer(start_text, parse_mode="Markdown", reply_markup=main_kb())

@router.message(Command("pin"))
async def cmd_pin(message: types.Message):
    args = message.text.split(maxsplit=1)
    if len(args) < 2:
        await message.answer("Сураныч, ПИН-кодду туура киргизиңиз. Мисалы: `/pin K7X9W2`", parse_mode="Markdown")
        return
    
    pin = args[1].strip()
    success, msg = await db.join_group_by_pin(message.from_user.id, pin)
    await message.answer(msg)

@router.message(Command("make_admin"))
async def cmd_make_admin(message: types.Message, state: FSMContext):
    await state.set_state(Form.create_group_name)
    await message.answer("Топтун (конкурстун) аталышын жазыңыз:", reply_markup=cancel_kb())

@router.message(Form.create_group_name)
async def process_group_name(message: types.Message, state: FSMContext):
    await state.update_data(group_name=message.text)
    await state.set_state(Form.create_group_cats)
    await message.answer("Эми конкурс үчүн категорияларды үтүр менен бөлүп жазыңыз (Мисалы: `КК, Рес, Тес`):", parse_mode="Markdown", reply_markup=cancel_kb())

@router.message(Form.create_group_cats)
async def process_group_cats(message: types.Message, state: FSMContext):
    cats = message.text.split(",")
    data = await state.get_data()
    pin = await db.create_group_with_categories(message.from_user.id, data['group_name'], cats)
    await state.clear()
    
    await message.answer(f"🎉 **Конкурс/Топ ийгиликтүү түзүлдү!**\n\n📌 ПИН-код: `{pin}`\nКатышуучуларга ушул кодду бериңиз.", parse_mode="Markdown", reply_markup=main_kb())

# --- АДМИН ПАНЕЛИ: ДАТА АРАЛЫГЫН ТАНДОО МЕНЮСУ ---
@router.message(Command("admin"))
async def cmd_admin_panel(message: types.Message):
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📅 Бүгүн", callback_data="admdate_today")],
        [InlineKeyboardButton(text="🗓 Ушул апта", callback_data="admdate_week")],
        [InlineKeyboardButton(text="📆 Ушул ай", callback_data="admdate_month")],
        [InlineKeyboardButton(text="♾ Бардык убакыт", callback_data="admdate_all")],
        [InlineKeyboardButton(text="✏️ Өзүм дата киргизем (Конкурс үчүн)", callback_data="admdate_custom")]
    ])
    await message.answer("📊 **кайсы дата аралыгындагы статистика керек?**", reply_markup=kb)

@router.callback_query(F.data.startswith("admdate_"))
async def process_admin_date_select(callback: types.CallbackQuery, state: FSMContext):
    choice = callback.data.split("_")[1]
    today = datetime.date.today()

    start_date = None
    end_date = None

    if choice == "today":
        start_date = today.strftime("%Y-%m-%d")
        end_date = start_date
    elif choice == "week":
        start_date = (today - datetime.timedelta(days=today.weekday())).strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")
    elif choice == "month":
        start_date = today.replace(day=1).strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")
    elif choice == "custom":
        await state.set_state(Form.custom_date_range)
        await callback.message.answer(
            "📅 Конкурстун датасын төмөнкү форматта жазыңыз:\n\n"
            "`КК.АА.ЖЖЖЖ - КК.АА.ЖЖЖЖ`\n\n"
            "Мисалы: `01.09.2026 - 15.09.2026`",
            parse_mode="Markdown",
            reply_markup=cancel_kb()
        )
        await callback.answer()
        return

    await generate_and_send_excel(callback.message, callback.from_user.id, start_date, end_date)
    await callback.answer()

@router.message(Form.custom_date_range)
async def process_custom_date_range(message: types.Message, state: FSMContext):
    try:
        parts = message.text.split("-")
        if len(parts) != 2:
            raise ValueError()

        d1 = datetime.datetime.strptime(parts[0].strip(), "%d.%m.%Y").strftime("%Y-%m-%d")
        d2 = datetime.datetime.strptime(parts[1].strip(), "%d.%m.%Y").strftime("%Y-%m-%d")

        await state.clear()
        await generate_and_send_excel(message, message.from_user.id, d1, d2)
    except Exception:
        await message.answer("⚠️ Ката формат! Сураныч, дал ушул үлгүдө киргизиңиз: `01.09.2026 - 15.09.2026`", parse_mode="Markdown", reply_markup=cancel_kb())

async def generate_and_send_excel(message: types.Message, admin_id: int, start_date: str = None, end_date: str = None):
    groups_data = await db.get_admin_excel_data_by_range(admin_id, start_date, end_date)
    if not groups_data:
        await message.answer("Сизде азырынча активдүү топтор жок же тандалган дата аралыгында окулган беттер табылган жок.", reply_markup=main_kb())
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for g_idx, group in enumerate(groups_data):
        ws = wb.create_sheet(title=f"Group_{g_idx+1}")
        
        range_str = f"{start_date} — {end_date}" if start_date and end_date else "Бардык убакыт"
        ws.append([f"👥 Топ: {group['name']}", f"PIN: {group['pin']}", f"📅 Аралык: {range_str}"])
        ws.append([])

        headers = ["Аты-жөнү", "Юзернейм"] + group['categories'] + ["Жалпы"]
        ws.append(headers)

        header_row_idx = 3
        total_cols = len(headers)

        for col_num in range(1, total_cols + 1):
            cell = ws.cell(row=header_row_idx, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in group['rows']:
            row_data = [row['name'], row['username']] + row['cats_pages'] + [row['total']]
            ws.append(row_data)

        max_row = header_row_idx + len(group['rows'])
        col_letter_max = openpyxl.utils.get_column_letter(total_cols)
        ws.auto_filter.ref = f"A{header_row_idx}:{col_letter_max}{max_row}"

        for r in range(header_row_idx, max_row + 1):
            for c in range(1, total_cols + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                if c > 2:
                    cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    document = BufferedInputFile(file_stream.read(), filename="contest_results.xlsx")
    await message.answer_document(document, caption="📊 **Конкурстун статистикалык Excel отчету даяр!**", reply_markup=main_kb())

# --- КАТЕГОРИЯЛАР МЕНЮСУ ---
@router.message(F.text == "📂 Категориялар")
async def show_categories(message: types.Message):
    cats = await db.get_user_categories(message.from_user.id)
    
    inline_keyboard = []
    for c_id, title, g_id in cats:
        prefix = "👥" if g_id else "👤"
        inline_keyboard.append([InlineKeyboardButton(text=f"{prefix} {title}", callback_data=f"cat_{c_id}")])
    
    inline_keyboard.append([InlineKeyboardButton(text="➕ Жаңы категория кошуу", callback_data="add_cat")])
    kb = InlineKeyboardMarkup(inline_keyboard=inline_keyboard)
    
    await message.answer("Категорияны тандаңыз же жаңысын кошуңуз:", reply_markup=kb)

@router.callback_query(F.data == "add_cat")
async def add_cat_callback(callback: types.CallbackQuery, state: FSMContext):
    await state.set_state(Form.add_category)
    await callback.message.answer("Жаңы жеке категориянын атын жазыңыз:", reply_markup=cancel_kb())
    await callback.answer()

@router.message(Form.add_category)
async def process_add_cat(message: types.Message, state: FSMContext):
    await db.create_user_category(message.from_user.id, message.text)
    await state.clear()
    await message.answer(f"✅ «{message.text}» категориясы кошулду!", reply_markup=main_kb())

@router.callback_query(F.data.startswith("cat_"))
async def select_category(callback: types.CallbackQuery, state: FSMContext):
    cat_id = int(callback.data.split("_")[1])
    await state.update_data(selected_cat_id=cat_id)
    await state.set_state(Form.input_pages)
    
    await callback.message.answer("Бүгүн канча бет окудуңуз? (Сан менен жазыңыз):", reply_markup=cancel_kb())
    await callback.answer()

# --- БЕТТИ КИРГИЗҮҮ ---
@router.message(Form.input_pages)
async def process_pages(message: types.Message, state: FSMContext):
    if not message.text.isdigit():
        await message.answer("Сураныч, сан менен гана жазыңыз:", reply_markup=cancel_kb())
        return
    
    pages = int(message.text)
    data = await state.get_data()
    cat_id = data['selected_cat_id']
    
    log_id = await db.log_pages(message.from_user.id, cat_id, pages)
    await state.clear()
    
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✏️ Өзгөртүү", callback_data=f"editlog_{log_id}"),
            InlineKeyboardButton(text="🗑 Өчүрүү", callback_data=f"dellog_{log_id}")
        ]
    ])
    
    await message.answer(f"🎉 Сонун! **{pages} бет** сакталды. 📚", parse_mode="Markdown", reply_markup=kb)
    await message.answer("Башкы меню:", reply_markup=main_kb())

@router.callback_query(F.data.startswith("dellog_"))
async def process_delete_log(callback: types.CallbackQuery):
    log_id = int(callback.data.split("_")[1])
    await db.delete_log(log_id)
    await callback.message.edit_text("❌ **Киргизилген беттер өчүрүлдү!**", parse_mode="Markdown")
    await callback.answer()

@router.callback_query(F.data.startswith("editlog_"))
async def process_start_edit_log(callback: types.CallbackQuery, state: FSMContext):
    log_id = int(callback.data.split("_")[1])
    await state.update_data(editing_log_id=log_id)
    await state.set_state(Form.edit_pages)
    await callback.message.answer("Беттин жаңы туура санын киргизиңиз:", reply_markup=cancel_kb())
    await callback.answer()

@router.message(Form.edit_pages)
async def process_save_edit_pages(message: types.Message, state: FSMContext):
    if not message.text.isdigit():
        await message.answer("Сураныч, сан менен гана жазыңыз:", reply_markup=cancel_kb())
        return
    
    new_pages = int(message.text)
    data = await state.get_data()
    log_id = data['editing_log_id']
    
    await db.update_last_log(log_id, new_pages)
    await state.clear()
    
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✏️ Өзгөртүү", callback_data=f"editlog_{log_id}"),
            InlineKeyboardButton(text="🗑 Өчүрүү", callback_data=f"dellog_{log_id}")
        ]
    ])
    
    await message.answer(f"✅ Жаңыртылды! **{new_pages} бет** болуп өзгөртүлдү. 📚", parse_mode="Markdown", reply_markup=kb)
    await message.answer("Башкы меню:", reply_markup=main_kb())

# --- ИНТЕРАКТИВДҮҮ СТАТИСТИКА ---
def build_month_keyboard(cats, year, month):
    keyboard = []
    for c_id, title, pages in cats:
        keyboard.append([
            InlineKeyboardButton(
                text=f"{title}: {pages}", 
                callback_data=f"stcat_{c_id}_{year}_{month}"
            )
        ])
    
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1

    keyboard.append([
        InlineKeyboardButton(text="<<", callback_data=f"stm_{prev_year}_{prev_month}"),
        InlineKeyboardButton(text=">>", callback_data=f"stm_{next_year}_{next_month}")
    ])
    return InlineKeyboardMarkup(inline_keyboard=keyboard)

@router.message(F.text == "📊 Статистика")
async def show_stats_main(message: types.Message):
    now = datetime.datetime.now()
    year, month = now.year, now.month
    
    cats, total = await db.get_monthly_category_summary(message.from_user.id, year, month)
    text = f"**{MONTH_NAMES[month]} {year}**\n**Pages: {total}**"
    kb = build_month_keyboard(cats, year, month)
    
    await message.answer(text, parse_mode="Markdown", reply_markup=kb)

@router.callback_query(F.data.startswith("stm_"))
async def process_stat_month_change(callback: types.CallbackQuery):
    _, y_str, m_str = callback.data.split("_")
    year, month = int(y_str), int(m_str)
    
    cats, total = await db.get_monthly_category_summary(callback.from_user.id, year, month)
    text = f"**{MONTH_NAMES[month]} {year}**\n**Pages: {total}**"
    kb = build_month_keyboard(cats, year, month)
    
    await callback.message.edit_text(text, parse_mode="Markdown", reply_markup=kb)
    await callback.answer()

@router.callback_query(F.data.startswith("stcat_"))
async def process_stat_cat_detail(callback: types.CallbackQuery):
    _, c_id_str, y_str, m_str = callback.data.split("_")
    c_id, year, month = int(c_id_str), int(y_str), int(m_str)
    
    cat_title, logs, total = await db.get_category_logs_by_month(callback.from_user.id, c_id, year, month)
    
    text = f"📂 **{cat_title} ({MONTH_NAMES[month]} {year}):**\n\n"
    if not logs:
        text += "Бул айда маалымат жок.\n"
    else:
        for date_str, pages in logs:
            text += f"📅 {date_str} — **{pages} бет**\n"
            
    text += f"\n💰 **Жалпы: {total}**"
    
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Back", callback_data=f"stm_{year}_{month}")]
    ])
    
    await callback.message.edit_text(text, parse_mode="Markdown", reply_markup=kb)
    await callback.answer()
